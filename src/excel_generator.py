"""Generate Excel reports with charts."""

from pathlib import Path
from typing import List, Dict
import logging

from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from .models import Workout, HeartRateZone

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports with multiple worksheets and charts."""

    def __init__(self, output_file: Path, theme_colors: Dict[str, str]):
        self.output_file = output_file
        self.theme_colors = theme_colors
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def _apply_header_style(self, sheet, row: int = 1):
        """Apply header styling to a row."""
        header_fill = PatternFill(
            start_color=self.theme_colors['header_bg'],
            end_color=self.theme_colors['header_bg'],
            fill_type='solid'
        )
        header_font = Font(
            color=self.theme_colors['header_text'],
            bold=True,
            size=11
        )

        for cell in sheet[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _apply_alternating_rows(self, sheet, start_row: int, end_row: int):
        """Apply alternating row colors."""
        alt_fill = PatternFill(
            start_color='F2F2F2',
            end_color='F2F2F2',
            fill_type='solid'
        )

        for row_idx in range(start_row, end_row + 1):
            if row_idx % 2 == 0:
                for cell in sheet[row_idx]:
                    if not cell.fill.start_color.rgb:  # Don't override headers
                        cell.fill = alt_fill

    def _autosize_columns(self, sheet):
        """Auto-size columns based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def create_summary_sheet(self, overall_stats: Dict):
        """Create Summary worksheet with high-level statistics."""
        logger.info("Creating Summary worksheet")

        sheet = self.workbook.create_sheet("Summary")

        # Title
        sheet['A1'] = "Apple Health Cycling Dashboard"
        sheet['A1'].font = Font(size=16, bold=True, color=self.theme_colors['primary'])

        # Overall statistics
        row = 3
        stats = [
            ("Total Rides", overall_stats['total_workouts'], ""),
            ("Total Distance", overall_stats['total_distance_km'], "km"),
            ("Total Time", overall_stats['total_duration_hours'], "hours"),
            ("Total Calories Burned", overall_stats['total_calories'], "kcal"),
            ("Total Elevation Gain", overall_stats['total_elevation_gain_m'], "m"),
            ("", "", ""),
            ("Average Distance per Ride", overall_stats['avg_distance_km'], "km"),
            ("Average Duration per Ride", overall_stats['avg_duration_minutes'], "min"),
            ("Average Speed", overall_stats['avg_speed_kmh'], "km/h"),
            ("Average Elevation Gain", overall_stats['avg_elevation_gain_m'], "m"),
            ("", "", ""),
            ("Longest Ride", overall_stats['max_distance_km'], "km"),
            ("Longest Duration", overall_stats['max_duration_minutes'], "min"),
            ("Biggest Climb", overall_stats['max_elevation_gain_m'], "m"),
            ("", "", ""),
            ("Years Active", overall_stats['years_active'], ""),
            ("First Ride", overall_stats['first_ride_date'], ""),
            ("Last Ride", overall_stats['last_ride_date'], ""),
        ]

        if overall_stats.get('avg_heart_rate'):
            stats.append(("Average Heart Rate", round(overall_stats['avg_heart_rate'], 1), "bpm"))

        for label, value, unit in stats:
            sheet[f'A{row}'] = label
            sheet[f'A{row}'].font = Font(bold=True)

            if isinstance(value, float):
                sheet[f'B{row}'] = round(value, 1)
            else:
                sheet[f'B{row}'] = value

            sheet[f'C{row}'] = unit
            row += 1

        self._autosize_columns(sheet)

    def create_yearly_trends_sheet(self, yearly_trends: List[Dict]):
        """Create Yearly Trends worksheet with charts."""
        logger.info("Creating Yearly Trends worksheet")

        sheet = self.workbook.create_sheet("Yearly Trends")

        # Convert to DataFrame for easy writing
        df = pd.DataFrame(yearly_trends)

        # Write headers
        headers = ['Year', 'Total Rides', 'Total Distance (km)', 'Total Time (hrs)',
                   'Total Calories', 'Total Elevation (m)', 'Avg Distance (km)',
                   'Avg Duration (min)', 'Avg Speed (km/h)', 'Avg Elevation (m)',
                   'Max Distance (km)', 'Max Elevation (m)']

        for col_idx, header in enumerate(headers, 1):
            sheet.cell(1, col_idx, header)

        # Write data
        for row_idx, row_data in enumerate(yearly_trends, 2):
            sheet.cell(row_idx, 1, row_data['year'])
            sheet.cell(row_idx, 2, row_data['total_workouts'])
            sheet.cell(row_idx, 3, round(row_data['total_distance_km'], 1))
            sheet.cell(row_idx, 4, round(row_data['total_duration_hours'], 1))
            sheet.cell(row_idx, 5, round(row_data['total_calories'], 0))
            sheet.cell(row_idx, 6, round(row_data['total_elevation_gain_m'], 0))
            sheet.cell(row_idx, 7, round(row_data['avg_distance_km'], 1))
            sheet.cell(row_idx, 8, round(row_data['avg_duration_minutes'], 1))
            sheet.cell(row_idx, 9, round(row_data['avg_speed_kmh'], 1))
            sheet.cell(row_idx, 10, round(row_data['avg_elevation_gain_m'], 0))
            sheet.cell(row_idx, 11, round(row_data['max_distance_km'], 1))
            sheet.cell(row_idx, 12, round(row_data['max_elevation_gain_m'], 0))

        self._apply_header_style(sheet, 1)
        self._apply_alternating_rows(sheet, 2, len(yearly_trends) + 1)
        self._autosize_columns(sheet)

        # Create charts
        self._create_yearly_distance_chart(sheet, len(yearly_trends))
        self._create_yearly_rides_chart(sheet, len(yearly_trends))

    def _create_yearly_distance_chart(self, sheet, num_years: int):
        """Create line chart for total distance by year."""
        chart = LineChart()
        chart.title = "Total Distance by Year"
        chart.style = 10
        chart.y_axis.title = "Distance (km)"
        chart.x_axis.title = "Year"

        data = Reference(sheet, min_col=3, min_row=1, max_row=num_years + 1)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=num_years + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, f"A{num_years + 4}")

    def _create_yearly_rides_chart(self, sheet, num_years: int):
        """Create bar chart for total rides by year."""
        chart = BarChart()
        chart.title = "Total Rides by Year"
        chart.style = 10
        chart.y_axis.title = "Number of Rides"
        chart.x_axis.title = "Year"

        data = Reference(sheet, min_col=2, min_row=1, max_row=num_years + 1)
        categories = Reference(sheet, min_col=1, min_row=2, max_row=num_years + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, f"J{num_years + 4}")

    def create_monthly_activity_sheet(self, monthly_trends: List[Dict]):
        """Create Monthly Activity worksheet."""
        logger.info("Creating Monthly Activity worksheet")

        sheet = self.workbook.create_sheet("Monthly Activity")

        # Write headers
        headers = ['Year', 'Month', 'Month Name', 'Total Rides', 'Total Distance (km)',
                   'Total Time (hrs)', 'Total Calories', 'Total Elevation (m)',
                   'Avg Distance (km)', 'Avg Speed (km/h)']

        for col_idx, header in enumerate(headers, 1):
            sheet.cell(1, col_idx, header)

        # Write data
        for row_idx, row_data in enumerate(monthly_trends, 2):
            sheet.cell(row_idx, 1, row_data['year'])
            sheet.cell(row_idx, 2, row_data['month'])
            sheet.cell(row_idx, 3, row_data['month_name'])
            sheet.cell(row_idx, 4, row_data['total_workouts'])
            sheet.cell(row_idx, 5, round(row_data['total_distance_km'], 1))
            sheet.cell(row_idx, 6, round(row_data['total_duration_hours'], 1))
            sheet.cell(row_idx, 7, round(row_data['total_calories'], 0))
            sheet.cell(row_idx, 8, round(row_data['total_elevation_gain_m'], 0))
            sheet.cell(row_idx, 9, round(row_data['avg_distance_km'], 1))
            sheet.cell(row_idx, 10, round(row_data['avg_speed_kmh'], 1))

        self._apply_header_style(sheet, 1)
        self._apply_alternating_rows(sheet, 2, len(monthly_trends) + 1)
        self._autosize_columns(sheet)

        # Create chart
        self._create_monthly_chart(sheet, len(monthly_trends))

    def _create_monthly_chart(self, sheet, num_months: int):
        """Create chart for monthly activity."""
        chart = LineChart()
        chart.title = "Monthly Distance Trend"
        chart.style = 10
        chart.y_axis.title = "Distance (km)"
        chart.x_axis.title = "Month"

        data = Reference(sheet, min_col=5, min_row=1, max_row=num_months + 1)
        categories = Reference(sheet, min_col=3, min_row=2, max_row=num_months + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Reduce number of x-axis labels if too many months
        if num_months > 24:
            chart.x_axis.tickLblSkip = num_months // 12

        sheet.add_chart(chart, f"A{num_months + 4}")

    def create_hr_analysis_sheet(self, hr_zones: Dict[str, HeartRateZone], workouts: List[Workout]):
        """Create Heart Rate Analysis worksheet."""
        logger.info("Creating Heart Rate Analysis worksheet")

        sheet = self.workbook.create_sheet("Heart Rate Analysis")

        # Heart Rate Zones section
        sheet['A1'] = "Heart Rate Zones"
        sheet['A1'].font = Font(size=14, bold=True)

        headers = ['Zone', 'BPM Range', 'Time (hours)', 'Time (minutes)']
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(3, col_idx, header)

        row = 4
        for zone_name, zone in hr_zones.items():
            sheet.cell(row, 1, zone_name)
            sheet.cell(row, 2, f"{zone.min_bpm}-{zone.max_bpm}")
            sheet.cell(row, 3, round(zone.time_hours, 1))
            sheet.cell(row, 4, round(zone.time_minutes, 1))
            row += 1

        self._apply_header_style(sheet, 3)
        self._apply_alternating_rows(sheet, 4, row - 1)

        # Average HR by year section
        sheet[f'A{row + 2}'] = "Average Heart Rate by Year"
        sheet[f'A{row + 2}'].font = Font(size=14, bold=True)

        # Calculate yearly avg HR
        from collections import defaultdict
        yearly_hr = defaultdict(list)

        for workout in workouts:
            if workout.avg_heart_rate:
                yearly_hr[workout.year].append(workout.avg_heart_rate)

        headers = ['Year', 'Avg Heart Rate (bpm)', 'Workouts with HR Data']
        header_row = row + 4
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(header_row, col_idx, header)

        data_row = header_row + 1
        for year in sorted(yearly_hr.keys()):
            hrs = yearly_hr[year]
            avg_hr = sum(hrs) / len(hrs)

            sheet.cell(data_row, 1, year)
            sheet.cell(data_row, 2, round(avg_hr, 1))
            sheet.cell(data_row, 3, len(hrs))
            data_row += 1

        self._apply_header_style(sheet, header_row)
        self._apply_alternating_rows(sheet, header_row + 1, data_row - 1)
        self._autosize_columns(sheet)

    def create_elevation_stats_sheet(self, yearly_trends: List[Dict], top_climbs: List[Workout]):
        """Create Elevation Stats worksheet."""
        logger.info("Creating Elevation Stats worksheet")

        sheet = self.workbook.create_sheet("Elevation Stats")

        # Yearly elevation section
        sheet['A1'] = "Elevation Gain by Year"
        sheet['A1'].font = Font(size=14, bold=True)

        headers = ['Year', 'Total Elevation (m)', 'Avg Elevation per Ride (m)', 'Max Single Climb (m)']
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(3, col_idx, header)

        row = 4
        for trend in yearly_trends:
            sheet.cell(row, 1, trend['year'])
            sheet.cell(row, 2, round(trend['total_elevation_gain_m'], 0))
            sheet.cell(row, 3, round(trend['avg_elevation_gain_m'], 0))
            sheet.cell(row, 4, round(trend['max_elevation_gain_m'], 0))
            row += 1

        self._apply_header_style(sheet, 3)
        self._apply_alternating_rows(sheet, 4, row - 1)

        # Top climbs section
        sheet[f'A{row + 2}'] = "Top 10 Biggest Climbs"
        sheet[f'A{row + 2}'].font = Font(size=14, bold=True)

        headers = ['Date', 'Distance (km)', 'Elevation Gain (m)', 'Duration (min)', 'Source']
        header_row = row + 4
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(header_row, col_idx, header)

        data_row = header_row + 1
        for workout in top_climbs:
            sheet.cell(data_row, 1, workout.date_str)
            sheet.cell(data_row, 2, round(workout.distance_km, 1))
            sheet.cell(data_row, 3, round(workout.elevation_gain_m or 0, 0))
            sheet.cell(data_row, 4, round(workout.duration_minutes, 0))
            sheet.cell(data_row, 5, workout.source_name)
            data_row += 1

        self._apply_header_style(sheet, header_row)
        self._apply_alternating_rows(sheet, header_row + 1, data_row - 1)
        self._autosize_columns(sheet)

        # Create chart
        self._create_elevation_chart(sheet, len(yearly_trends))

    def _create_elevation_chart(self, sheet, num_years: int):
        """Create bar chart for elevation gain by year."""
        chart = BarChart()
        chart.title = "Total Elevation Gain by Year"
        chart.style = 11
        chart.y_axis.title = "Elevation (m)"
        chart.x_axis.title = "Year"

        data = Reference(sheet, min_col=2, min_row=3, max_row=num_years + 3)
        categories = Reference(sheet, min_col=1, min_row=4, max_row=num_years + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, "G3")

    def create_raw_data_sheet(self, workouts: List[Workout]):
        """Create Raw Data worksheet with all workouts."""
        logger.info("Creating Raw Data worksheet")

        sheet = self.workbook.create_sheet("Raw Data")

        # Headers
        headers = ['Date', 'Year', 'Month', 'Distance (km)', 'Duration (min)',
                   'Avg Speed (km/h)', 'Calories', 'Elevation Gain (m)',
                   'Avg HR (bpm)', 'Max HR (bpm)', 'Source', 'GPX File']

        for col_idx, header in enumerate(headers, 1):
            sheet.cell(1, col_idx, header)

        # Write workout data
        for row_idx, workout in enumerate(sorted(workouts, key=lambda w: w.start_date), 2):
            sheet.cell(row_idx, 1, workout.date_str)
            sheet.cell(row_idx, 2, workout.year)
            sheet.cell(row_idx, 3, workout.month)
            sheet.cell(row_idx, 4, round(workout.distance_km, 2))
            sheet.cell(row_idx, 5, round(workout.duration_minutes, 1))
            sheet.cell(row_idx, 6, round(workout.avg_speed_kmh, 1))
            sheet.cell(row_idx, 7, round(workout.calories, 0))
            sheet.cell(row_idx, 8, round(workout.elevation_gain_m or 0, 0))
            sheet.cell(row_idx, 9, round(workout.avg_heart_rate, 1) if workout.avg_heart_rate else "")
            sheet.cell(row_idx, 10, round(workout.max_heart_rate, 1) if workout.max_heart_rate else "")
            sheet.cell(row_idx, 11, workout.source_name)
            sheet.cell(row_idx, 12, workout.gpx_file or "")

        self._apply_header_style(sheet, 1)
        self._apply_alternating_rows(sheet, 2, len(workouts) + 1)
        self._autosize_columns(sheet)

    def save(self):
        """Save the workbook to file."""
        logger.info(f"Saving Excel report to {self.output_file}")
        self.workbook.save(self.output_file)
        logger.info(f"Report saved successfully")
